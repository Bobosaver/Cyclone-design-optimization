import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.image as mpimg  
from scipy.optimize import curve_fit
import numpy as np
import openpyxl as opxl
import copy
from scipy.interpolate import griddata
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import interpn

# 功能1:旋风分离器响应面拟合
def spline_interpolation():
    file = st.file_uploader("请选择上传文件", type="xlsx")   #上传文件

    if file is not None:
        data = opxl.load_workbook(file)
        table = data['Sheet1']
        max_row = table.max_row
        st.write(max_row)

        columnA = table['A2':'A%d' % max_row]  #提取风速
        Fengsu = []
        for column_cells in columnA:
            for cell in column_cells:
                Fengsu.append(cell.value)

        columnB = table['B2':'B%d' % max_row]  # 提取颗粒粒径
        Keli = []
        for column_cells in columnB:
            for cell in column_cells:
                Keli.append(cell.value)

        columnD = table['D2':'D%d' % max_row]  #提取逃逸率
        escape = []
        for column_cells in columnD:
            for cell in column_cells:
                escape.append(cell.value)
                
        x = np.asarray(Fengsu)  #数据预处理
        y = np.asarray(Keli)
        z = 1- np.asarray(escape)
        x_1 = list(set(x))
        y_1 = list(set(y))
        points = (np.array(x_1), np.array(y_1))
        z = z.reshape(len(x_1),len(y_1))
        xi = np.arange(5,24,0.1)
        yi = np.arange(1,30,0.1)
        xi, yi = np.meshgrid(xi, yi)

        zi = interpn(points,z,(xi,yi), method='splinef2d' )  # 样条插值拟合

        from mpl_toolkits.mplot3d import Axes3D  # 绘制拟合响应面
        fig = plt.figure()
        plt.figure(figsize=(10,10),dpi=175)
        ax = fig.add_subplot(111, projection='3d')
        ax.scatter(x, y, 100*z,s=5,color='red',)
        ax.plot_surface(xi, yi, 100*zi,cmap='rainbow')
        ax.set_facecolor("white")
        ax.xaxis.set_tick_params(color='r',labelsize=10)
        ax.yaxis.set_tick_params(color='r',labelsize=10)
        ax.zaxis.set_tick_params(color='r',labelsize=10)
        ax.set_xlabel('wind velocity(m/s)',size=10,labelpad=5)
        ax.set_ylabel('Particle size(μm)',size=10,labelpad=5)
        ax.set_zlabel('Collection rate(%)',size=10,labelpad=5)
        ax.set_title('Response surface',size=10) 
        ax.view_init(32, -32)
        st.pyplot(fig)
        
    else:
        st.text('还未选择文件！')

# 功能2:旋风分离器收集率曲线拟合
def collection_curve():
    file = st.file_uploader("请选择上传文件", type="xlsx")   #上传文件
    m = st.sidebar.slider("请输入粒径(μm):",1,15,30)

    if file is not None:
        data = opxl.load_workbook(file)  # 提取原始数据
        table = data['Sheet1']
        max_row = table.max_row

        columnA = table['A2':'A%d' % max_row]  # 提取颗粒粒径
        Keli = []
        for column_cells in columnA:
            for cell in column_cells:
                Keli.append(cell.value)

        columnB = table['B2':'B%d' % max_row]  # 提取收集率
        escape = []
        for column_cells in columnB:
            for cell in column_cells:
                escape.append(cell.value)

        y = np.asarray(Keli)  # 数据预处理
        z = 1 - np.asarray(escape)
       
        def func(x, a, b):   #非线性双R曲线拟合
            return np.exp(-(x/a)**b)
        ab, para = curve_fit(func,y,z)
        
        #计算拟合优度
        mean = np.mean(z)  
        ss_tot = np.sum((z - mean) ** 2)  
        ss_res = np.sum((z - func(y, *ab)) ** 2)
        r_squared = 1 - (ss_res / ss_tot)  

        #绘图
        fig = plt.figure()
        plt.scatter(y, 100*(1-z),s=20, marker='.',edgecolors='red')
        yi = np.arange(1,30,0.1 )
        Z = func(yi,*ab)
        plt.plot(yi, 100*(1-Z), lw=2)
        plt.yticks(fontsize=15,color='#000000')
        plt.xticks(fontsize=15,color='#000000')
        plt.xlabel('Particle size(μm)',fontsize=15)
        plt.ylabel('Collection rate(%)',fontsize=15)
        st.pyplot(fig)
        st.text('注：红点为仿真计算值，蓝色曲线为拟合双R曲线')
        a = ab[0]
        b = ab[1]
        shoujilv = 1-np.exp(-(m/a)**b)
        
        st.sidebar.write("<span style='color: red;'>收集率为：</span>", unsafe_allow_html=True)
        st.sidebar.write(str(shoujilv))        
        st.sidebar.write("<span style='color: red;'>拟合优度为：</span>", unsafe_allow_html=True)
        st.sidebar.write(str(r_squared))
    else:
        st.text('还未选择文件！')

def Tcrc():
    file_1 = st.file_uploader("请选择上传分析拟合文件", type="xlsx")   # 上传文件
    file_2 = st.file_uploader("请选择上传实际颗粒粒径分布数据", type="xlsx")   # 上传文件
    v = st.sidebar.slider("请指定当前风速(m/s):",5,17,24)

    if file_1 and file_2 is not None:
        data1 = opxl.load_workbook(file_1)  # 读取仿真分析数据
        table1 = data1['Sheet1']
        max_row = table1.max_row
        columnA = table1['A2':'A%d' % max_row]  # 提取风速
        Fengsu = []
        for column_cells in columnA:
            for cell in column_cells:
                Fengsu.append(cell.value)
        columnB = table1['B2':'B%d' % max_row]  # 提取颗粒粒径
        Keli = []
        for column_cells in columnB:
            for cell in column_cells:
                Keli.append(cell.value)
        columnD = table1['D2':'D%d' % max_row]  # 提取逃逸率
        escape = []
        for column_cells in columnD:
            for cell in column_cells:
                escape.append(cell.value)
        x = np.asarray(Fengsu)  # 数据预处理
        y = np.asarray(Keli)
        z = 1- np.asarray(escape)
        xi = np.linspace(5,17,1000)
        yi = np.linspace(1,30,1000)
        xi, yi = np.meshgrid(xi, yi)
        
        zi = griddata((x, y), z, (xi, yi), method='cubic')  # 进行响应面拟合

        Measured_keli = opxl.load_workbook(file_2)
        sheet2 = Measured_keli['Sheet1']
        max_row2 = sheet2.max_row
        # 提取实际颗粒粒径与密度分布
        columnA = sheet2['A2':'A%d' % max_row2]
        M_Keli = []
        for column_cells in columnA:
            for cell in column_cells:
                M_Keli.append(cell.value)
                
        columnC = sheet2['C2':'C%d' % max_row2]
        Density = []
        for column_cells in columnC:
            for cell in column_cells:
               Density.append(cell.value)
               
        for i in range(len(M_Keli)):  # 去除可能空值
            if M_Keli[i]==None:
                M_Keli.remove(None)
            if Density[i]==None:
                Density.remove(None)
               
        #分割超微粒径与超大粒径
        mini_keli=[]
        max_keli=[]
        MK = copy.deepcopy(M_Keli)
        for i in range(len(M_Keli)):
            a = M_Keli[i]
            if a <30:
                MK[i]=(30*a)**0.5
            if a >30:
                max_keli.append(a)
                MK.remove(a)

        #拟合各粒径收集率
        Collection_rate_mid = griddata((x, y), z, (v,MK), method='cubic')
        Collection_rate_max = [1 for _ in range(len(max_keli))] 
        Collection_rate = [*Collection_rate_mid,*Collection_rate_max] 

        #计算总收集效率
        T_rate = 0
        for i in range(len(Collection_rate)):
            R = Density[i]*Collection_rate[i]
            T_rate = T_rate + R
            
        st.sidebar.write("<span style='color: red;'>颗粒总收集率为：</span>", unsafe_allow_html=True)
        st.sidebar.write(str(T_rate))
    
    
#主程序功能介绍
st.title('旋风分离器参数拟合')
st.header('现有功能')

#可选功能
op_func =['旋风分离器响应面拟合','旋风分离器收集率曲线拟合及收集效率计算','旋风分离器总收集效率计算']

#选择执行功能
choose_func = st.radio('请选择需要执行的功能',op_func,index=None)

if choose_func == op_func[0]:
    st.write("<p style='color: black; text-indent:2em; line-height:30px;'>该功能为旋风分离器颗粒收集率的响应面拟合：\
    旋风分离器包含多个结构设计参数及边界条件，例如筒高及入口风速，对颗粒收集率有不同的影响水平。通过对多参数设计下的颗粒收集率\
    进行算法拟合,可以得到响应面函数，有助于进行进一步设计优化。</p>",unsafe_allow_html=True)
    st.write("<p style='color: black; text-indent:2em; line-height:30px;'>为使用该功能，需要指定一个Excel文件，包含多组\
    数据样本，具体的数据格式参考Help中指南。</p>",unsafe_allow_html=True)
    spline_interpolation()
    
elif choose_func == op_func[1]:
    st.write("<p style='color: black; text-indent:2em; line-height:30px;'>该功能为旋风分离器的颗粒收集率曲线的拟合：在\
    旋风分离器的结构及边界条件确定时，不同粒径下的颗粒收集率呈类双指数函数分布。本功能以Rosin-Rammler曲线为拟合目标曲线，通过\
    非线性最小二乘法进行拟合，以预测颗粒粒径与收集率间的关系。</p>",unsafe_allow_html=True)
    st.write("<p style='color: black; text-indent:2em; line-height:30px;'>为使用该功能，需要指定一个Excel文件，包含不同\
    粒径下的颗粒收集率数据样本，具体的数据格式参考Help中指南。</p>",unsafe_allow_html=True)
    collection_curve()
    
elif choose_func == op_func[2]:
    Tcrc()

